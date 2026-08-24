# -*- coding: utf-8 -*-
import json
import os
import tempfile
import threading
import unittest
from urllib.error import HTTPError
from urllib.request import Request, urlopen

from openpyxl import Workbook, load_workbook

import sys

_BASE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
if _BASE not in sys.path:
    sys.path.insert(0, _BASE)

import panel_server
from nota_pendiente import CABECERA_TAREAS


class TestPanelServer(unittest.TestCase):

    def setUp(self):
        self.temporal = tempfile.TemporaryDirectory()
        self.obra = 'OBRA PRUEBA'
        carpeta_obra = os.path.join(self.temporal.name, self.obra)
        os.makedirs(carpeta_obra)
        self.ruta_xlsx = os.path.join(carpeta_obra, 'FICHA DE OBRA.xlsx')

        wb = Workbook()
        wb.active.title = 'Datos'
        ws = wb.create_sheet('Tareas')
        ws.append(CABECERA_TAREAS)
        ws.append([
            'Revisar cuadro', 'Parte de obra', '22/08/2026',
            'parte-22-08.pdf', 'Pendiente',
        ])
        wb.save(self.ruta_xlsx)
        wb.close()

        self.servidor = panel_server.crear_servidor(
            puerto=0, directorio_obras=self.temporal.name)
        self.hilo = threading.Thread(
            target=self.servidor.serve_forever, daemon=True)
        self.hilo.start()
        host, puerto = self.servidor.server_address[:2]
        self.assertEqual(host, '127.0.0.1')
        self.url = f'http://127.0.0.1:{puerto}/api/marcar_hecho'

    def tearDown(self):
        self.servidor.shutdown()
        self.servidor.server_close()
        self.hilo.join(timeout=3)
        self.temporal.cleanup()

    def _post(self, **cambios):
        datos = {
            'obra_carpeta': self.obra,
            'tarea': 'Revisar cuadro',
            'origen': 'Parte de obra',
            'fecha': '22/08/2026',
            'archivo': 'parte-22-08.pdf',
        }
        datos.update(cambios)
        peticion = Request(
            self.url,
            data=json.dumps(datos).encode('utf-8'),
            headers={'Content-Type': 'application/json'},
            method='POST',
        )
        try:
            respuesta = urlopen(peticion, timeout=3)
        except HTTPError as error:
            return error.code, json.loads(error.read().decode('utf-8'))
        with respuesta:
            return respuesta.status, json.loads(
                respuesta.read().decode('utf-8'))

    def test_post_real_marca_la_tarea_y_responde_200(self):
        estado_http, cuerpo = self._post()

        wb = load_workbook(self.ruta_xlsx, data_only=False)
        estado = wb['Tareas']['E2'].value
        wb.close()

        self.assertEqual(estado_http, 200)
        self.assertTrue(cuerpo['ok'])
        self.assertEqual(estado, 'Hecho')

    def test_fila_inexistente_responde_404_sin_modificar_el_fichero(self):
        with open(self.ruta_xlsx, 'rb') as fichero:
            contenido_antes = fichero.read()

        estado_http, cuerpo = self._post(tarea='Tarea inexistente')

        with open(self.ruta_xlsx, 'rb') as fichero:
            contenido_despues = fichero.read()

        self.assertEqual(estado_http, 404)
        self.assertFalse(cuerpo['ok'])
        self.assertEqual(contenido_despues, contenido_antes)

    def test_rechaza_obra_con_punto_punto(self):
        estado_http, cuerpo = self._post(obra_carpeta='../OBRA PRUEBA')

        self.assertEqual(estado_http, 400)
        self.assertFalse(cuerpo['ok'])

    def test_rechaza_ruta_relativa_a_unidad_de_windows(self):
        """Una ruta "relativa a unidad" (p.ej. 'Z:evil') no es absoluta ni
        contiene '..', pero en Windows os.path.join descarta el directorio
        base cuando la letra de unidad no coincide (ntpath.join). Se elige
        una unidad distinta a la del propio directorio temporal para que la
        fuga se reproduzca sin depender de en qué disco viva %TEMP%."""
        unidad_base, _ = os.path.splitdrive(self.temporal.name)
        otra_unidad = 'Z' if unidad_base[:1].upper() != 'Z' else 'Y'

        estado_http, cuerpo = self._post(
            obra_carpeta=f'{otra_unidad}:evil')

        self.assertEqual(estado_http, 400)
        self.assertFalse(cuerpo['ok'])


if __name__ == '__main__':
    unittest.main()
