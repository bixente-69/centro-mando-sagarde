# -*- coding: utf-8 -*-
import os
import sys
import tempfile
import unittest

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font

_BASE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
if _BASE not in sys.path:
    sys.path.insert(0, _BASE)

from nota_pendiente import (
    CABECERA_TAREAS,
    anadir_tarea_pendiente,
    marcar_tarea_hecha,
)


class TestNotaPendiente(unittest.TestCase):

    def _crear_ficha(self, carpeta, preparar=None):
        ruta = os.path.join(carpeta, 'FICHA DE OBRA.xlsx')
        wb = Workbook()
        wb.active.title = 'Datos'
        if preparar:
            preparar(wb)
        wb.save(ruta)
        wb.close()
        return ruta

    def _anadir(self, ruta):
        anadir_tarea_pendiente(
            ruta,
            tarea='Revisar temas del correo de obra',
            origen='Correo reenviado por Iker',
            fecha='20/08/2026',
            archivo='TEMAS PENDIENTES 20-08-2026.txt',
        )

    def test_crea_hoja_con_cabecera_y_fila_si_no_existe(self):
        with tempfile.TemporaryDirectory() as carpeta:
            ruta = self._crear_ficha(carpeta)
            self._anadir(ruta)
            wb = load_workbook(ruta, data_only=False)
            filas = list(wb['Tareas'].iter_rows(values_only=True))
            wb.close()

        self.assertEqual(filas, [
            tuple(CABECERA_TAREAS),
            ('Revisar temas del correo de obra',
             'Correo reenviado por Iker', '20/08/2026',
             'TEMAS PENDIENTES 20-08-2026.txt', 'Pendiente'),
        ])

    def test_anade_al_final_sin_tocar_filas_existentes(self):
        def preparar(wb):
            ws = wb.create_sheet('Tareas')
            ws.append(CABECERA_TAREAS)
            ws.append([
                'Tarea anterior', 'WhatsApp', '19/08/2026',
                'NOTA ANTERIOR.txt', 'Hecho',
            ])

        with tempfile.TemporaryDirectory() as carpeta:
            ruta = self._crear_ficha(carpeta, preparar)
            self._anadir(ruta)
            wb = load_workbook(ruta, data_only=False)
            filas = list(wb['Tareas'].iter_rows(values_only=True))
            wb.close()

        self.assertEqual(filas[0], tuple(CABECERA_TAREAS))
        self.assertEqual(filas[1], (
            'Tarea anterior', 'WhatsApp', '19/08/2026',
            'NOTA ANTERIOR.txt', 'Hecho'))
        self.assertEqual(filas[2], (
            'Revisar temas del correo de obra',
            'Correo reenviado por Iker', '20/08/2026',
            'TEMAS PENDIENTES 20-08-2026.txt', 'Pendiente'))

    def test_otra_hoja_con_datos_formula_y_estilo_sobrevive(self):
        def preparar(wb):
            ws = wb.create_sheet('Riesgos')
            ws.append(['Riesgo', 'Estado', 'Control'])
            ws.append(['Retraso de suministro', 'Abierto', '=1+1'])
            ws['A2'].font = Font(bold=True)

        with tempfile.TemporaryDirectory() as carpeta:
            ruta = self._crear_ficha(carpeta, preparar)
            self._anadir(ruta)
            wb = load_workbook(ruta, data_only=False)
            riesgos = list(wb['Riesgos'].iter_rows(values_only=True))
            formula = wb['Riesgos']['C2'].value
            negrita = wb['Riesgos']['A2'].font.bold
            wb.close()

        self.assertEqual(riesgos, [
            ('Riesgo', 'Estado', 'Control'),
            ('Retraso de suministro', 'Abierto', '=1+1'),
        ])
        self.assertEqual(formula, '=1+1')
        self.assertTrue(negrita)

    def test_marca_la_fila_pendiente_exacta_sin_tocar_las_demas(self):
        def preparar(wb):
            ws = wb.create_sheet('Tareas')
            ws.append(CABECERA_TAREAS)
            ws.append([
                'Revisar cuadro', 'Parte de obra', '22/08/2026',
                'parte-22-08.pdf', 'Pendiente',
            ])
            ws.append([
                'Revisar cuadro distinto', 'Parte de obra', '22/08/2026',
                'parte-22-08.pdf', 'Pendiente',
            ])
            riesgos = wb.create_sheet('Riesgos')
            riesgos.append(['Riesgo', 'Control'])
            riesgos.append(['Suministro', '=1+1'])

        with tempfile.TemporaryDirectory() as carpeta:
            ruta = self._crear_ficha(carpeta, preparar)
            resultado = marcar_tarea_hecha(
                ruta, tarea='Revisar cuadro', origen='Parte de obra',
                fecha='22/08/2026', archivo='parte-22-08.pdf')
            wb = load_workbook(ruta, data_only=False)
            filas = list(wb['Tareas'].iter_rows(values_only=True))
            riesgos = list(wb['Riesgos'].iter_rows(values_only=True))
            wb.close()

        self.assertTrue(resultado)
        self.assertEqual(filas, [
            tuple(CABECERA_TAREAS),
            ('Revisar cuadro', 'Parte de obra', '22/08/2026',
             'parte-22-08.pdf', 'Hecho'),
            ('Revisar cuadro distinto', 'Parte de obra', '22/08/2026',
             'parte-22-08.pdf', 'Pendiente'),
        ])
        self.assertEqual(riesgos, [
            ('Riesgo', 'Control'),
            ('Suministro', '=1+1'),
        ])

    def test_fila_no_encontrada_devuelve_false_sin_modificar_el_fichero(self):
        def preparar(wb):
            ws = wb.create_sheet('Tareas')
            ws.append(CABECERA_TAREAS)
            ws.append([
                'Revisar cuadro', 'Parte de obra', '22/08/2026',
                'parte-22-08.pdf', 'Pendiente',
            ])

        with tempfile.TemporaryDirectory() as carpeta:
            ruta = self._crear_ficha(carpeta, preparar)
            with open(ruta, 'rb') as fichero:
                contenido_antes = fichero.read()

            resultado = marcar_tarea_hecha(
                ruta, tarea='Otra tarea', origen='Parte de obra',
                fecha='22/08/2026', archivo='parte-22-08.pdf')

            with open(ruta, 'rb') as fichero:
                contenido_despues = fichero.read()

        self.assertFalse(resultado)
        self.assertEqual(contenido_despues, contenido_antes)

    def test_sin_hoja_tareas_devuelve_false_sin_crearla(self):
        with tempfile.TemporaryDirectory() as carpeta:
            ruta = self._crear_ficha(carpeta)
            with open(ruta, 'rb') as fichero:
                contenido_antes = fichero.read()

            resultado = marcar_tarea_hecha(
                ruta, tarea='Revisar cuadro', origen='Parte de obra',
                fecha='22/08/2026', archivo='parte-22-08.pdf')

            with open(ruta, 'rb') as fichero:
                contenido_despues = fichero.read()
            wb = load_workbook(ruta, data_only=False)
            hojas = wb.sheetnames
            wb.close()

        self.assertFalse(resultado)
        self.assertEqual(hojas, ['Datos'])
        self.assertEqual(contenido_despues, contenido_antes)

    def test_dos_filas_iguales_solo_cambia_la_que_esta_pendiente(self):
        def preparar(wb):
            ws = wb.create_sheet('Tareas')
            ws.append(CABECERA_TAREAS)
            ws.append([
                'Revisar cuadro', 'Parte de obra', '22/08/2026',
                'parte-22-08.pdf', 'pEnDiEnTe',
            ])
            ws.append([
                'Revisar cuadro', 'Parte de obra', '22/08/2026',
                'parte-22-08.pdf', 'HECHO',
            ])

        with tempfile.TemporaryDirectory() as carpeta:
            ruta = self._crear_ficha(carpeta, preparar)
            resultado = marcar_tarea_hecha(
                ruta, tarea='Revisar cuadro', origen='Parte de obra',
                fecha='22/08/2026', archivo='parte-22-08.pdf')
            wb = load_workbook(ruta, data_only=False)
            estados = [wb['Tareas'].cell(fila, 5).value for fila in (2, 3)]
            wb.close()

        self.assertTrue(resultado)
        self.assertEqual(estados, ['Hecho', 'HECHO'])


if __name__ == '__main__':
    unittest.main()
