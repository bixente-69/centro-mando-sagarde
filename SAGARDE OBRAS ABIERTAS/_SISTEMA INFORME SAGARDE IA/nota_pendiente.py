# -*- coding: utf-8 -*-
"""Añade una tarea pendiente a la ficha Excel de una obra."""
import argparse

from openpyxl import load_workbook


CABECERA_TAREAS = ('Tarea', 'Origen', 'Fecha', 'Archivo', 'Estado')


def anadir_tarea_pendiente(ruta_xlsx, tarea, origen, fecha, archivo):
    """Añade una fila a ``Tareas`` sin modificar las demás hojas o filas."""
    wb = load_workbook(ruta_xlsx, data_only=False)
    try:
        if 'Tareas' in wb.sheetnames:
            ws = wb['Tareas']
        else:
            ws = wb.create_sheet('Tareas')
            ws.append(CABECERA_TAREAS)

        ws.append([tarea, origen, fecha, archivo, 'Pendiente'])
        wb.save(ruta_xlsx)
    finally:
        wb.close()


def marcar_tarea_hecha(ruta_xlsx, tarea, origen, fecha, archivo):
    """Marca la primera tarea pendiente que coincide exactamente."""
    wb = load_workbook(ruta_xlsx, data_only=False)
    try:
        if 'Tareas' not in wb.sheetnames:
            return False

        ws = wb['Tareas']
        columnas = {
            celda.value: celda.column
            for celda in ws[1]
            if celda.value in CABECERA_TAREAS
        }
        if any(cabecera not in columnas for cabecera in CABECERA_TAREAS):
            return False

        valores_buscados = (tarea, origen, fecha, archivo)
        for numero_fila in range(2, ws.max_row + 1):
            valores_fila = tuple(
                ws.cell(numero_fila, columnas[cabecera]).value
                for cabecera in CABECERA_TAREAS[:4]
            )
            estado = ws.cell(numero_fila, columnas['Estado']).value
            if (valores_fila == valores_buscados
                    and str(estado).casefold() == 'pendiente'):
                ws.cell(numero_fila, columnas['Estado']).value = 'Hecho'
                wb.save(ruta_xlsx)
                return True
        return False
    finally:
        wb.close()


def main():
    parser = argparse.ArgumentParser(
        description='Añade una tarea pendiente a FICHA DE OBRA.xlsx.')
    parser.add_argument('ruta_xlsx', help='Ruta de FICHA DE OBRA.xlsx')
    parser.add_argument('--tarea', required=True)
    parser.add_argument('--origen', required=True)
    parser.add_argument('--fecha', required=True, help='Fecha DD/MM/AAAA')
    parser.add_argument('--archivo', required=True,
                        help='Nombre del fichero de nota asociado')
    args = parser.parse_args()

    anadir_tarea_pendiente(
        args.ruta_xlsx, tarea=args.tarea, origen=args.origen,
        fecha=args.fecha, archivo=args.archivo)
    print(f"Tarea pendiente añadida a: {args.ruta_xlsx}")


if __name__ == '__main__':
    main()
