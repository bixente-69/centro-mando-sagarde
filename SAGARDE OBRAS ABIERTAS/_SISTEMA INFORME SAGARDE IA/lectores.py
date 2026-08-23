# -*- coding: utf-8 -*-
"""
LECTORES AUXILIARES (capa 1, genericos)
----------------------------------------
- leer_ficha(ruta_xlsx): lee la FICHA DE OBRA.xlsx (Datos/Personal/Hitos/
  Riesgos/Plan semanal/Tareas) a un dict. Tolera hojas ausentes.
- leer_materiales(ruta_xlsx): lee la hoja de entrega de materiales de forma
  DEFENSIVA. El formato real es desordenado (categorias como filas
  cabecera, items repetidos, columnas de fecha variables, columna TOTAL al
  final). Devuelve, por cada pestana-mes, los items con su total, y detecta
  la fecha de ultima entrada para avisar si esta desactualizada.
- listar_documentos(carpeta): recorre la carpeta de la obra y clasifica
  cada archivo por tipo, con ruta relativa para enlazar.

Ninguno de estos lectores asume que el material esta instalado: solo lee lo
que hay en la hoja. La interpretacion (recibido vs instalado) queda a criterio
del encargado, como pide el proyecto.
"""
import os
from datetime import datetime, date

try:
    import openpyxl
except ImportError:
    openpyxl = None


def leer_ficha(ruta_xlsx):
    ficha = {'datos': {}, 'personal': [], 'hitos': [], 'riesgos': [],
             'plan': [], 'tareas': []}
    if not openpyxl or not os.path.isfile(ruta_xlsx):
        ficha['_disponible'] = False
        return ficha
    ficha['_disponible'] = True
    wb = openpyxl.load_workbook(ruta_xlsx, data_only=True)

    if 'Datos' in wb.sheetnames:
        ws = wb['Datos']
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row and row[0]:
                clave = str(row[0]).strip()
                valor = '' if len(row) < 2 or row[1] is None else str(row[1]).strip()
                ficha['datos'][clave] = valor

    def _tabla(nombre, dest):
        if nombre not in wb.sheetnames:
            return
        ws = wb[nombre]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return
        cab = [str(c).strip() if c is not None else '' for c in rows[0]]
        for row in rows[1:]:
            if not row or all(c is None or str(c).strip() == '' for c in row):
                continue
            reg = {}
            vacio_util = True
            for i, c in enumerate(cab):
                v = row[i] if i < len(row) else None
                v = '' if v is None else (v.strftime('%d/%m/%Y') if isinstance(v, (datetime, date)) else str(v).strip())
                if c:
                    reg[c] = v
                if c and v and not v.lower().startswith('(ejemplo'):
                    vacio_util = True
            # descartar filas de ejemplo
            primero = str(row[0]).strip().lower() if row[0] else ''
            if primero.startswith('(ejemplo'):
                continue
            dest.append(reg)

    _tabla('Personal', ficha['personal'])
    _tabla('Hitos', ficha['hitos'])
    _tabla('Riesgos', ficha['riesgos'])
    _tabla('Plan semanal', ficha['plan'])
    _tabla('Tareas', ficha['tareas'])
    return ficha


def leer_materiales(ruta_xlsx):
    """
    Devuelve dict:
      {
        'disponible': bool,
        'meses': ['Febrero26', ...],
        'ultimo_mes': 'Mayo26',
        'ultima_fecha': '08/05/2026' | None,
        'dias_desde': int | None,
        'items': [ {'categoria','material','tipo','uni','total'}, ... ],  # del ultimo mes
        'aviso': str | None,
      }
    """
    res = {'disponible': False, 'meses': [], 'ultimo_mes': None, 'ultima_fecha': None,
           'dias_desde': None, 'items': [], 'aviso': None}
    if not openpyxl or not os.path.isfile(ruta_xlsx):
        return res
    res['disponible'] = True
    wb = openpyxl.load_workbook(ruta_xlsx, data_only=True)
    res['meses'] = list(wb.sheetnames)
    if not res['meses']:
        return res
    ultimo = res['meses'][-1]
    res['ultimo_mes'] = ultimo
    ws = wb[ultimo]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return res
    header = rows[0]

    # localizar columna TOTAL (o T.SUMINISTRADO)
    total_idx = None
    for i, h in enumerate(header):
        if isinstance(h, str) and h.strip().upper() in ('TOTAL', 'T.SUMINISTRADO'):
            total_idx = i
            break

    # ultima fecha con datos (cabeceras que son fechas)
    fechas = [h for h in header if isinstance(h, (datetime, date))]
    if fechas:
        f = max(fechas)
        res['ultima_fecha'] = f.strftime('%d/%m/%Y')
        res['dias_desde'] = (datetime.now().date() - (f.date() if isinstance(f, datetime) else f)).days

    if total_idx is None:
        res['aviso'] = "No se localizó la columna TOTAL en la hoja; se muestra solo la referencia al archivo."
        return res

    cat = None
    for r in rows[1:]:
        name = r[0]
        if not name or not isinstance(name, str):
            continue
        name = name.strip()
        tipo = r[1] if len(r) > 1 else None
        data_cols = r[4:total_idx]
        tot = r[total_idx] if total_idx < len(r) else None
        has_data = any(x not in (None, '') for x in data_cols) or (tot not in (None, '', 0))
        if not has_data and (tipo in (None, '')):
            cat = name  # fila cabecera de categoria
            continue
        if tot in (None, '', 0):
            continue
        res['items'].append({
            'categoria': cat or '', 'material': name,
            'tipo': '' if tipo is None else str(tipo),
            'uni': '' if len(r) < 4 or r[3] is None else str(r[3]),
            'total': tot,
        })

    if res['dias_desde'] is not None and res['dias_desde'] > 30:
        res['aviso'] = (f"La hoja de materiales no se actualiza desde {res['ultima_fecha']} "
                        f"({res['dias_desde']} días). Los datos de consumo/stock pueden estar desfasados.")
    return res


TIPOS_DOC = {
    'Planos y gráfica': ('.dwg', '.dxf'),
    'PDF': ('.pdf',),
    'Word': ('.docx', '.doc'),
    'Excel': ('.xlsx', '.xls', '.xlsm'),
    'Imágenes': ('.jpg', '.jpeg', '.png', '.gif', '.bmp', '.heic'),
    'Comprimidos': ('.zip', '.rar', '.7z'),
}


def _clasifica(ext):
    ext = ext.lower()
    for cat, exts in TIPOS_DOC.items():
        if ext in exts:
            return cat
    return 'Otros'


def listar_documentos(carpeta_obra, base_href_desde):
    """
    Recorre la carpeta de la obra y devuelve lista de documentos con ruta
    relativa (href) desde la ubicacion del panel HTML.
    Ignora la propia carpeta de informes generados.
    """
    docs = []
    if not os.path.isdir(carpeta_obra):
        return docs
    for root, dirs, files in os.walk(carpeta_obra):
        dirs[:] = [d for d in dirs if d != 'INFORME SAGARDE IA']
        for fn in files:
            if fn.startswith('~$') or fn.lower() in ('plot.log',):
                continue
            ruta_abs = os.path.join(root, fn)
            ext = os.path.splitext(fn)[1]
            rel = os.path.relpath(ruta_abs, base_href_desde).replace('\\', '/')
            subcarpeta = os.path.relpath(root, carpeta_obra).replace('\\', '/')
            if subcarpeta == '.':
                subcarpeta = ''
            try:
                size = os.path.getsize(ruta_abs)
            except OSError:
                size = 0
            docs.append({
                'nombre': fn, 'categoria': _clasifica(ext),
                'subcarpeta': subcarpeta, 'href': rel,
                'kb': round(size / 1024),
            })
    docs.sort(key=lambda d: (d['categoria'], d['subcarpeta'], d['nombre']))
    return docs
